from __future__ import annotations

import logging
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from document_automation_studio.engine.rule_engine import RuleEngine
from document_automation_studio.models.rule_models import RuleSet, TextReplacementRule

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Processor for Excel documents with rule-driven automation."""

    def __init__(self) -> None:
        self.logger = logger
        self.rule_engine = RuleEngine()

    def process(
        self,
        source_path: Path,
        output_root: Path,
        preserve_folder_structure: bool = True,
        input_root: Path | None = None,
        rule_set: RuleSet | None = None,
    ) -> Path:
        """Process an .xlsx file and save the result into the output root."""
        if source_path.suffix.lower() != ".xlsx":
            raise ValueError("Unsupported file type for ExcelProcessor: %s" % source_path)

        workbook = load_workbook(source_path)
        self.logger.debug("Processing Excel workbook %s", source_path)

        replacements = rule_set.excel_replacements if rule_set and rule_set.excel_replacements else (rule_set.text_replacements if rule_set else [])
        if replacements:
            self._replace_text(workbook, replacements)

        if rule_set:
            self._apply_insert_rows(workbook, rule_set.excel_insert_rows)
            self._apply_insert_columns(workbook, rule_set.excel_insert_columns)
            self._apply_hyperlinks(workbook, rule_set.excel_hyperlinks)

        destination = self._build_destination(source_path, output_root, preserve_folder_structure, input_root)
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
        self.logger.info("Excel workbook saved to %s", destination)
        return destination

    def _replace_text(self, workbook: object, replacements: Iterable[TextReplacementRule]) -> None:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=False):
                for cell in row:
                    if isinstance(cell.value, str):
                        new_value = self.rule_engine.apply_text_replacements(cell.value, replacements)
                        if new_value != cell.value:
                            self.logger.debug("Replacing Excel cell %s in sheet %s", cell.coordinate, sheet.title)
                            cell.value = new_value

    def _apply_insert_rows(self, workbook: object, row_rules: list[dict[str, object]]) -> None:
        for rule in row_rules:
            sheet_name = rule.get("sheet")
            index = int(rule.get("index", 1))
            values = rule.get("values", [])
            sheet = self._find_sheet(workbook, sheet_name)
            if sheet is None:
                continue
            sheet.insert_rows(index)
            if isinstance(values, list):
                for col_index, value in enumerate(values, start=1):
                    sheet.cell(row=index, column=col_index, value=value)

    def _apply_insert_columns(self, workbook: object, column_rules: list[dict[str, object]]) -> None:
        for rule in column_rules:
            sheet_name = rule.get("sheet")
            index = int(rule.get("index", 1))
            values = rule.get("values", [])
            sheet = self._find_sheet(workbook, sheet_name)
            if sheet is None:
                continue
            sheet.insert_cols(index)
            if isinstance(values, list):
                for row_index, value in enumerate(values, start=1):
                    sheet.cell(row=row_index, column=index, value=value)

    def _apply_hyperlinks(self, workbook: object, hyperlink_rules: list[dict[str, str]]) -> None:
        for rule in hyperlink_rules:
            sheet_name = rule.get("sheet")
            target = rule.get("cell")
            url = rule.get("url")
            display = rule.get("display", url)
            sheet = self._find_sheet(workbook, sheet_name)
            if sheet is None or not target or not url:
                continue
            sheet[target].hyperlink = url
            sheet[target].value = display

    def _find_sheet(self, workbook: object, sheet_name: str | None) -> Worksheet | None:
        if not sheet_name:
            return workbook.active
        return workbook[sheet_name] if sheet_name in workbook.sheetnames else None

    def _build_destination(
        self,
        source_path: Path,
        output_root: Path,
        preserve_folder_structure: bool,
        input_root: Path | None,
    ) -> Path:
        if preserve_folder_structure and input_root is not None:
            relative = source_path.relative_to(input_root)
            return output_root / relative
        return output_root / source_path.name
