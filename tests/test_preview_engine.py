from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from document_automation_studio.engine.preview_engine import PreviewEngine
from document_automation_studio.engine.rule_engine import RuleEngine
from document_automation_studio.models.rule_models import RuleSet, TextReplacementRule


def test_preview_engine_excel_replacement(tmp_path: Path) -> None:
    source = tmp_path / "workbook.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Company: {{Company}}"
    workbook.save(source)

    engine = PreviewEngine(RuleEngine())
    rule_set = RuleSet(excel_replacements=[TextReplacementRule(find="{{Company}}", replace="ACME")])

    preview_items = engine.preview_files(
        [source],
        rule_set=rule_set,
        input_root=tmp_path,
        max_items=10,
        process_word_files=False,
        process_excel_files=True,
    )

    assert len(preview_items) == 1
    assert preview_items[0].change_type == "Excel Replacement"
    assert preview_items[0].old_value == "Company: {{Company}}"
    assert preview_items[0].new_value == "Company: ACME"


def test_preview_engine_skips_excel_when_disabled(tmp_path: Path) -> None:
    source = tmp_path / "workbook.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Company: {{Company}}"
    workbook.save(source)

    engine = PreviewEngine(RuleEngine())
    rule_set = RuleSet(excel_replacements=[TextReplacementRule(find="{{Company}}", replace="ACME")])

    preview_items = engine.preview_files(
        [source],
        rule_set=rule_set,
        input_root=tmp_path,
        max_items=10,
        process_word_files=False,
        process_excel_files=False,
    )

    assert len(preview_items) == 0


def test_preview_engine_excel_rename_preview(tmp_path: Path) -> None:
    source = tmp_path / "workbook.xlsx"
    workbook = Workbook()
    workbook.save(source)

    engine = PreviewEngine(RuleEngine())
    rule_set = RuleSet(rename_patterns={"workbook": "workbook_final"})

    preview_items = engine.preview_files([source], rule_set=rule_set, input_root=tmp_path, max_items=10)

    assert len(preview_items) == 1
    assert preview_items[0].change_type == "File Rename"
    assert preview_items[0].old_value == "workbook.xlsx"
    assert preview_items[0].new_value == "workbook_final.xlsx"
