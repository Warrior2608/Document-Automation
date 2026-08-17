from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from document_automation_studio.models.rule_models import RuleSet, TextReplacementRule
from document_automation_studio.processors.excel_processor import ExcelProcessor


def test_excel_processor_replaces_text(tmp_path: Path) -> None:
    source = tmp_path / "sample.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Company: {{Company}}"
    workbook.save(source)

    processor = ExcelProcessor()
    output_root = tmp_path / "out"
    output_root.mkdir(exist_ok=True)

    rule_set = RuleSet(text_replacements=[TextReplacementRule(find="{{Company}}", replace="ACME")])
    destination = processor.process(
        source_path=source,
        output_root=output_root,
        preserve_folder_structure=False,
        rule_set=rule_set,
    )

    assert destination.exists()
    from openpyxl import load_workbook

    loaded = load_workbook(destination)
    assert loaded.active["A1"].value == "Company: ACME"
